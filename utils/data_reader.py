import openpyxl
import json

def load_excel_data(filepath):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    data = []
    for row in sheet.iter_rows(min_row=2):
        entry = {}
        for i, cell in enumerate(row):
            header = headers[i]
            if header:
                value = cell.value
                entry[header] = "" if value is None else str(value).strip()
        data.append(entry)

    return data


def load_json_data(filepath):
    with open(filepath, "r", encoding="utf-8") as file:
        return json.load(file)